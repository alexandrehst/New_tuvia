import json
import logging
import random
import uuid
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import range_boundaries, get_column_letter

from controller.ctrl_planos import Ctrl_Planos
from dto.LP_Perguntas_dto import LP_Perguntas_dto
from model.LP_Tipo_Plano import LP_Tipo_Plano
from model.User import User
from notification.evento import EventoFactory
from service.BubbleAPIService import BubbleAPIService
from service.PlanGenerator import PlanGenerator
from service.PlanService import PlanService
import re
import os

from service.brevo import BrevoService

logger = logging.getLogger()

class Ctrl_Lp_Planos:
    #init

    plan_generator = PlanGenerator(model='gpt-4o', temperature=1, max_tokens=4096)

    def __init__(self):
        pass

    def get_perguntas(self, tipo_plano_id, tipo_plano_descricao):

        resposta = BubbleAPIService().get_generic_by_constraint(BubbleAPIService().PERGUNTAS_PLANO, [{'key': 'Tipo_Plano', 'constraint_type': 'equals', 'value': tipo_plano_id}])

        if resposta['error']:
            logger.error('Erro ao recuperar perguntas')
            return None

        if resposta[BubbleAPIService().PERGUNTAS_PLANO]:
            return resposta[BubbleAPIService().PERGUNTAS_PLANO]
        else:
            resposta = self.make_respostas( tipo_plano_id, tipo_plano_descricao)
            if resposta['error']:
                logger.error('Erro ao recuperar perguntas')
                return None

        perguntas_json = resposta['message']

        dto = LP_Perguntas_dto( BubbleAPIService() )
        try:
            tipo_plano = LP_Tipo_Plano(_id=tipo_plano_id, Descricao= tipo_plano_descricao)
            perguntas = [dto._map({'pergunta': pergunta, 'tipo_plano': tipo_plano}) for pergunta in perguntas_json['Questions']]
        except Exception as e:
            logger.error(f'Erro ao mapear perguntas: {e}')
            return None

        dto.insert_bulk(perguntas)

        resposta = BubbleAPIService().get_generic_by_constraint(BubbleAPIService().PERGUNTAS_PLANO, [{'key': 'Tipo_Plano', 'constraint_type': 'equals', 'value': tipo_plano_id}])

        if resposta['error']:
            logger.error('Erro ao recuperar perguntas')
            return None

        if resposta[BubbleAPIService().PERGUNTAS_PLANO]:
            return resposta[BubbleAPIService().PERGUNTAS_PLANO]


    def gerar(self, user, empresa, tema_plano, perguntas, respostas):
        plan = self.make_plan( tema_plano, perguntas, respostas)

        if not plan:
            return False

        plano_de_trabalho = self.make_plano_de_trabalho(plan)

        if not plano_de_trabalho:
            logger.error('Erro ao gerar plano de trabalho')
            return False

        filename = self.make_xls(plan, plano_de_trabalho, empresa, user)

        brevo = BrevoService()

        resposta = brevo.create_user(user)

        evento = EventoFactory().create_evento(EventoFactory().LP_PLANEJAMENTO, autor=user)
        evento.execute( user, filename)

        return filename


    def make_respostas(self, tipo_plano_id, tipo_plano_descricao):
        logger.info(f"Make plan LP.")

        value_system = f"Seu papel é me auxiliar em um planejamento estratégico. Me faça cinco perguntas, necessárias para a criação de objetivos. "
        value_user = f"O tema do planejamento é {tipo_plano_descricao}."

        messages = [
            {'role': 'system', 'content': value_system},
            {'role': 'user', 'content': value_user}
        ]

        resp =  self.plan_generator.generate_perguntas(messages)
        return resp

    def make_plan(self, tema_plano, perguntas, respostas):

        logger.info(f"Make plan LP.")

        value_system = f"Seu papel é me auxiliar em um planejamento estratégico para {tema_plano}\nAbaixo estão algumas perguntas e respostas. "
        value_user = ""
        for pergunta, resposta in zip(perguntas, respostas):
            value_user += f"Pergunta: {pergunta}\nResposta: {resposta}\n"

        value_user += "Garanta que a lista tenha no mínimo cinco objetivos, cada um com pelo menos três resultados-chave"

        messages = [
            {'role': 'system', 'content': value_system},
            {'role': 'user', 'content': value_user}
        ]

        chatGPT_answer = self.plan_generator.generate_plan(messages)

        try:
            if chatGPT_answer['error']:
                return None

            plan = PlanService.create_from_answer(title=tema_plano, business_info=None, improvement_needed=None, success_indicator=None, resposta=chatGPT_answer['message'])

            if plan is None:
                logger.error(f"Erro na criação do plano")
                return None

            while len(plan.okrs) < 5:
                logger.info(f"Criando objetivos adicionais")
                okr = Ctrl_Planos.make_objective_adicional(messages, plan.okrs)
                if okr is None:
                   break
                plan.okrs.append(okr)

            # Checa se todos os objetivos tem 3 resultados chave e adiciona se necessário
            for okr in plan.okrs:
                while len(okr.key_results) < 3:
                    logger.info(f"Criando resultados chave adicionais")
                    kr = Ctrl_Planos.make_key_result(id, okr)
                    if kr is None:
                        break
                    okr.key_results.append(kr)
            return plan

        except Exception as e:
            logger.error(f"Erro na geração do plano: {e}")
            return None

    def make_plano_de_trabalho(self, plan):
        logger.info(f"Make plano de trabalho LP.")

        value_system = f"Seu papel é me auxiliar na criação de um plano de atividade para o plano estratégico que será fornecido pelo usuário. "
        value_system += "O plano deve conter atividades para três meses e que levem a um impacto nos valores dos resultados chaves. "
        value_system += "Como vou usar esse resultado em um documento, preciso que você não inclua introdução ou conclusão. Apenas o plano de trabalho."
        value_system += "Utilize os seguintes simbolos de markdown para os níveis hierárquicos do plano: #, ##, ###, ####."

        value_user = f"Plano: {plan.title}\n"

        for okr in plan.okrs:
            value_user += f"OKR: {okr.titulo}\n"
            for key_result in okr.key_results:
                value_user += f"Resultado chave: {key_result.descricao}\n"
                value_user += f"Meta: {key_result.tipo_metrica} {key_result.value} {key_result.unit}\n"

        messages = [
            {'role': 'system', 'content': value_system},
            {'role': 'user', 'content': value_user}
        ]

        chatGPT_answer = self.plan_generator.model.generate(messages)

        if chatGPT_answer['error']:
            logger.error(f"Erro na geração do plano de trabalho - ChatGPT: {chatGPT_answer['message']}")
            return None

        return chatGPT_answer['message']

    def make_xls(self, plan, plano_de_trabalho, empresa, user):

        # Cria um novo workbook e ativa a primeira planilha
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Configurações básicas
        sheet.title = f"PE - {empresa}"
        fonte_titulo = Font(bold=True, size=18, color='0b5394')
        fonte_objetivo = Font(bold=True, size=14, color='666666')
        font_resultado_chave_titulo = Font(size=9, color='000000')
        font_resultado_numero = Font(size=12, color='FFFFFF')

        font_titulo1 = Font(bold=True, size=14, color='434343')
        font_titulo2 = Font(italic=True, size=12, color='5c5c5c')
        font_atividade = Font(size=12, color='5c5c5c')
        font_pequena = Font(size=11, color='7f7f7f')

        obj_fill = PatternFill(start_color='efefef', end_color='efefef', fill_type='solid')
        obj_0_fill = PatternFill(start_color='8db3e2', end_color='8db3e2', fill_type='solid')
        obj_1_fill = PatternFill(start_color='548dd4', end_color='548dd4', fill_type='solid')
        obj_2_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        obj_3_fill = PatternFill(start_color='17365d', end_color='17365d', fill_type='solid')
        obj_4_fill = PatternFill(start_color='0f243e', end_color='17365d', fill_type='solid')
        cores_objetivos = [obj_0_fill, obj_1_fill, obj_2_fill, obj_3_fill, obj_4_fill, obj_0_fill, obj_1_fill, obj_2_fill, obj_3_fill, obj_4_fill]

        fonte_plano_1 = Font(bold=True, size=14, color='FFFFFF')
        fill_plano_1 = PatternFill(start_color='0070c0', end_color='0070c0', fill_type='solid')
        fonte_plano_2 = Font(bold=True, size=13, color='666666')
        fill_plano_2 = PatternFill(start_color='efefef', end_color='efefef', fill_type='solid')
        fonte_plano_3 = Font(size=12, color='FFFFFF')
        fill_plano_3 = PatternFill(start_color='434343', end_color='434343', fill_type='solid')
        fonte_plano_4 = Font(size=10, color='5c5c5c')
        fill_plano_4 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        fontes_plano = [fonte_titulo, fonte_plano_1, fonte_plano_2, fonte_plano_3, fonte_plano_4]
        fills_plano = [fill_plano_4,fill_plano_1, fill_plano_2, fill_plano_3, fill_plano_4]


        proxima_cor = iter(cores_objetivos)
        light_gray = PatternFill(start_color='bfbfbf', end_color='bfbfbf', fill_type='solid')
        dark_gray = PatternFill(start_color='434343', end_color='434343', fill_type='solid')

        center_alignment = Alignment(horizontal='center', vertical='center')

        sheet.sheet_view.showGridLines = False

        # Set the width of columns
        sheet.column_dimensions['A'].width = 1
        sheet.column_dimensions['B'].width = 8
        max_descricao_length = max(len(resultado_chave.descricao) for objetivo in plan.okrs for resultado_chave in objetivo.key_results)
        sheet.column_dimensions['C'].width = max_descricao_length

        sheet.row_dimensions[1].height = 8.25

        sheet.append([""])

        sheet.append(["", f'Plano para: {plan.title}'])
        sheet.append(["", f"Feito por: {user.nome}"])
        sheet.append(["", "Gerado pela Tuvia"])
        self.bordas(sheet, 'B2:F4')

        sheet.append([""])
        sheet.append(["", plan.title])
        sheet.append(["", 'Objetivos, indicadores e metas para acompanhar o avanço do plano.'])
        sheet.append([""])

        sheet.merge_cells('B6:F6')
        sheet['B6'].font = fonte_titulo
        #sheet['B6'].alignment = center_alignment
        sheet['B7'].font = font_pequena

        # Inclui objetivos
        linha = 9
        for objetivo in plan.okrs:
            sheet.append(["", objetivo.titulo])
            sheet.append(["", "Resultados-chave:"])
            sheet.append(['','#', 'Descrição',	'Unidade', 'Valor alvo', 'Valor atual'])
            for i, resultado_chave in enumerate(objetivo.key_results):
                sheet.append(['', i+1, resultado_chave.descricao, resultado_chave.unit, resultado_chave.value, ''])
            sheet.append([""])

            # Titulos dos objetivos
            cells = 'B' + str(linha) + ':F' + str(linha)
            sheet.merge_cells( cells)
            top_left_cell =  'B' + str(linha)  # Obtém a célula superior esquerda
            sheet[top_left_cell].fill = obj_fill
            sheet[top_left_cell].font = fonte_objetivo

            self.bordas(sheet, f'B{linha}:F{linha+5}')

            # Linha "resultados chave"
            linha += 1
            cells = 'B' + str(linha) + ':F' + str(linha)
            sheet.merge_cells( cells)
            top_left_cell =  'B' + str(linha) # Obtém a célula superior esquerda
            sheet[top_left_cell].fill = light_gray
            sheet[top_left_cell].font = font_resultado_chave_titulo

            # linhas do resultado chave
            linha += 1
            for cell in ['B' + str(linha), 'C' + str(linha), 'D' + str(linha), 'E' + str(linha), 'F' + str(linha)]:
                sheet[cell].fill = dark_gray
                sheet[cell].font = font_resultado_chave_titulo
                sheet[cell].alignment = center_alignment

            # numeros dos objetivos
            cor = next(proxima_cor)
            for i in range(3):
                linha += 1
                sheet['B' + str(linha)].font = font_resultado_numero
                sheet['B' + str(linha)].fill = cor
                sheet['B' + str(linha)].alignment = center_alignment
                sheet['D' + str(linha)].alignment = center_alignment
                sheet['E' + str(linha)].alignment = center_alignment

            linha += 2


        # Checa se o plano de trabalho possui título
        primeira_linha = plano_de_trabalho.split('\n')[0]
        if "Mês" not in primeira_linha:
            texto_primeira_linha = self.remove_non_letter(primeira_linha)
        else:
            texto_primeira_linha = "Plano de Trabalho"
            # Inclui o plano de trabalho
            sheet.append(['', "Plano de Trabalho"])
            sheet.append([""])
            linha += 2

        sheet.append([""])
        linha += 1
        sheet.append(["", texto_primeira_linha])
        cells = 'B' + str(linha) + ':F' + str(linha)
        sheet.merge_cells(cells)
        sheet['B' + str(linha)].font = fonte_titulo
        sheet.append(['', 'Plano para a execução do plano estratégico, incluindo etapas, meses e entregáveis.'])
        linha += 1
        sheet['B' + str(linha)].font = font_pequena
        linha += 1

        markdowns = self.extrai_markdowns(plano_de_trabalho)

        linha_inicio_borda = 0
        linha_fim_borda = 0
        for line in plano_de_trabalho.split('\n'):
            nivel = self.nivel_markdown(line, markdowns)
            if nivel == 0:
                continue # não imprime o titulo

            sheet.append(['', self.remove_non_letter(line)])

            # logica para aplicar da borda ao redor do nível 2
            if nivel <= 2 and linha_fim_borda != 0: # fim da borda
                linha_fim_borda = linha -2
                self.bordas(sheet, f'B{linha_inicio_borda}:F{linha_fim_borda}')
                linha_inicio_borda = 0
                linha_fim_borda = 0

            if nivel == 2 and linha_inicio_borda == 0: # início da borda
               linha_inicio_borda = linha
               linha_fim_borda = linha # serve como marcação de que está dentro da borda

            sheet['B' + str(linha)].font = fontes_plano[nivel]
            for cell in ['B' + str(linha), 'C' + str(linha), 'D' + str(linha), 'E' + str(linha), 'F' + str(linha)]:
                sheet[cell].fill = fills_plano[nivel]

            linha += 1

        #aplica a borda no final do plano
        self.bordas(sheet, f'B{linha_inicio_borda}:F{linha}')

        # Salva o workbook
        random_number = random.randint(1000, 9999)
        filename = f"plano_para_{empresa}_gerado_pela_Tuvia_{random_number}.xlsx"

        current_path = os.getcwd()
        file_path = os.path.join(current_path, filename)
        workbook.save(file_path)

        logger.info(f"Planilha {filename} criada com sucesso!")
        return filename

    def bordas( self, ws, range_string):
        min_col, min_row, max_col, max_row = range_boundaries(range_string)

        # Convertendo índices de coluna para letras
        def col_idx_to_letter(col_idx):
            return openpyxl.utils.get_column_letter(col_idx)

        # Extraindo as bordas
        top_border = [f"{col_idx_to_letter(col)}{min_row}" for col in range(min_col, max_col + 1)]
        bottom_border = [f"{col_idx_to_letter(col)}{max_row}" for col in range(min_col, max_col + 1)]
        left_border = [f"{col_idx_to_letter(min_col)}{row}" for row in range(min_row, max_row + 1)]
        right_border = [f"{col_idx_to_letter(max_col)}{row}" for row in range(min_row, max_row + 1)]

        left_border = left_border[1:-1]
        right_border = right_border[1:-1]
        thin_side = Side(style='thin')

        # Corners
        first_cell = top_border[0]
        last_cell = top_border[-1]
        ws[first_cell].border = Border(top=thin_side, left=thin_side)
        ws[last_cell].border = Border(top=thin_side, right=thin_side)
        top_border = top_border[1:-1]

        first_cell = bottom_border[0]
        last_cell = bottom_border[-1]
        ws[first_cell].border = Border(bottom=thin_side, left=thin_side)
        ws[last_cell].border = Border(bottom=thin_side, right=thin_side)
        bottom_border = bottom_border[1:-1]

        # Aplicar borda superior
        for cell in top_border:
            ws[cell].border = Border(top=thin_side)

        # Aplicar borda inferior
        for cell in bottom_border:
            ws[cell].border = Border(bottom=thin_side)

        # Aplicar borda esquerda
        for cell in left_border:
            ws[cell].border = Border(left=thin_side)

        # Aplicar borda direita
        for cell in right_border:
            ws[cell].border = Border(right=thin_side)

    def extrai_markdowns(self, plano):
        #Detectar markdown utilizado na geração do plano de trabalho
        nivel = 1
        markdowns = []
        for line in plano.split('\n'):
            if re.match(r'^\W+', line):
                extracted_chars = re.match(r'^\W+', line).group()
                markdowns.append(extracted_chars)
                nivel += 1
                if nivel > 4: break

        return markdowns

    def nivel_markdown(self, line, markdowns):
        pattern = re.match(r'^\W+', line)
        extracted_chars = None
        if pattern:
            extracted_chars = pattern.group()

        nivel = 4
        for i, markdown in enumerate(markdowns):
            if extracted_chars == markdown:
                nivel = i
                break

        return nivel

    def remove_non_letter(self, text):
        return re.sub(r'[*#]', '', text.lstrip())
